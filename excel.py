import calendar
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from config import DEPARTMENTS_FULL_LINE, DOCTOR_NAME, MONTH_NAMES_PL
from db import Entry

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

COLUMNS = [
    "Dzień", "Oddział", "Godz. od", "Godz. do", "Godziny",
    "Dyżury od", "Dyżury do", "Dyżury godz.", "Razem godz.",
]
COL_DZIEN, COL_ODDZIAL, COL_GODZ_OD, COL_GODZ_DO, COL_GODZINY = 1, 2, 3, 4, 5
COL_DYZ_OD, COL_DYZ_DO, COL_DYZ_GODZ, COL_RAZEM = 6, 7, 8, 9

HEADER_ROW = 4


def _hhmm(dt: datetime) -> str:
    return dt.strftime("%H:%M")


def _aggregate(day_entries: list[Entry], kind: str):
    """Combine same-kind entries within one calendar day into a single (od, do, godziny, oddzial)
    row, matching the template's one-row-per-day layout. Godziny is the sum of each entry's
    already-rounded duration, mirroring how the paper template rounds per row."""
    kind_entries = [e for e in day_entries if e.kind == kind]
    if not kind_entries:
        return None, None, 0.0, None
    start = min(e.start_dt for e in kind_entries)
    end = max(e.end_dt for e in kind_entries)
    hours = round(sum(e.hours for e in kind_entries), 2)
    oddzial = kind_entries[0].oddzial
    return start, end, hours, oddzial


def _precise_total_hours(entries: list[Entry]) -> float:
    """Grand total computed from full-precision (unrounded) durations, then rounded once.
    Matches the reference template: summing the per-row *rounded* cells instead drifts by a
    cent or two over a full month (~30 independently-rounded numbers) versus this method."""
    total_minutes = sum((e.end_dt - e.start_dt).total_seconds() / 60 for e in entries)
    return round(total_minutes / 60, 2)


def generate_month_excel(
    entries: list[Entry], year: int, month: int, out_path: Path, doctor_name: str = DOCTOR_NAME,
) -> Path:
    days_in_month = calendar.monthrange(year, month)[1]

    by_day: dict[int, list[Entry]] = {d: [] for d in range(1, days_in_month + 1)}
    for e in entries:
        day = e.start_dt.day
        if day in by_day:
            by_day[day].append(e)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month:02d}.{year}"

    ncols = len(COLUMNS)
    last_col_letter = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "GRAFIK DYŻURÓW – LICZENIE GODZIN (Godziny + Dyżury = Razem)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = CENTER

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Lekar {doctor_name}    Miesiąc: {MONTH_NAMES_PL[month]}    Rok: {year}"
    ws["A2"].font = Font(bold=True)

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = f"Oddział {DEPARTMENTS_FULL_LINE}"

    for idx, title in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=title)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = BORDER

    first_data_row = HEADER_ROW + 1
    row = first_data_row
    for day in range(1, days_in_month + 1):
        day_entries = by_day[day]
        w_start, w_end, w_hours, w_odd = _aggregate(day_entries, "work")
        d_start, d_end, d_hours, d_odd = _aggregate(day_entries, "dyzur")
        oddzial = w_odd or d_odd or ""

        values = {
            COL_DZIEN: day,
            COL_ODDZIAL: oddzial,
            COL_GODZ_OD: _hhmm(w_start) if w_start else None,
            COL_GODZ_DO: _hhmm(w_end) if w_end else None,
            COL_GODZINY: w_hours,
            COL_DYZ_OD: _hhmm(d_start) if d_start else None,
            COL_DYZ_DO: _hhmm(d_end) if d_end else None,
            COL_DYZ_GODZ: d_hours,
        }
        for col_idx, val in values.items():
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = CENTER
            if col_idx in (COL_GODZINY, COL_DYZ_GODZ):
                cell.number_format = "0.00"

        razem_cell = ws.cell(row=row, column=COL_RAZEM)
        razem_cell.value = f"=E{row}+H{row}"
        razem_cell.number_format = "0.00"
        razem_cell.border = BORDER
        razem_cell.alignment = CENTER
        row += 1

    last_data_row = row - 1

    sum_row = row + 1
    ws.merge_cells(f"A{sum_row}:{get_column_letter(COL_RAZEM - 1)}{sum_row}")
    label_cell = ws.cell(row=sum_row, column=1, value="SUMA GODZIN (miesiąc):")
    label_cell.font = Font(bold=True)
    label_cell.border = BORDER

    total_cell = ws.cell(row=sum_row, column=COL_RAZEM)
    total_cell.value = _precise_total_hours(entries)
    total_cell.number_format = "0.00"
    total_cell.font = Font(bold=True)
    total_cell.border = BORDER
    total_cell.alignment = CENTER

    widths = [6, 20, 10, 10, 10, 10, 10, 12, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
