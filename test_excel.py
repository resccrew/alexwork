from datetime import datetime

import openpyxl
import pytest

import db
from config import TZ
from excel import HEADER_ROW, generate_month_excel

YEAR, MONTH = 2026, 3

# (day, kind, oddzial, start "HH:MM", end "HH:MM", end_day_offset) -- transcribed from the
# doctor's March 2026 paper "GRAFIK DYŻURÓW" template photo.
RAW_ENTRIES = [
    (2, "work", "Urologia", "06:56", "15:10", 0),
    (3, "work", "Urologia", "07:02", "16:18", 0),
    (4, "work", "Urologia", "07:20", "15:31", 0),
    (5, "work", "Urologia", "07:15", "15:22", 0),
    (6, "work", "Chirurgia Onk.", "07:08", "15:00", 0),
    (6, "dyzur", "Chirurgia Onk.", "15:00", "08:11", 1),
    (8, "dyzur", "Urologia", "16:46", "06:46", 1),
    (9, "work", "Urologia", "06:46", "16:03", 0),
    (10, "work", "Urologia", "07:00", "15:00", 0),
    (10, "dyzur", "Urologia", "15:00", "07:00", 1),
    (11, "work", "Urologia", "07:00", "10:51", 0),
    (12, "work", "Chirurgia Onk.", "07:14", "15:00", 0),
    (12, "dyzur", "Chirurgia Onk.", "15:00", "07:00", 1),
    (13, "work", "Urologia", "07:00", "17:11", 0),
    (14, "dyzur", "Chirurgia Ogólna", "08:28", "00:00", 1),
    (15, "dyzur", "Chirurgia Ogólna", "00:00", "08:46", 0),
    (16, "work", "Urologia", "07:06", "15:00", 0),
    (16, "dyzur", "Urologia", "15:00", "07:00", 1),
    (17, "work", "Urologia", "07:00", "09:46", 0),
    (18, "work", "Urologia", "07:15", "15:19", 0),
    (19, "work", "Chirurgia Onk.", "07:17", "15:00", 0),
    (19, "dyzur", "Chirurgia Onk.", "15:00", "07:00", 1),
    (20, "work", "Urologia", "07:00", "10:07", 0),
    (22, "dyzur", "Chirurgia Ogólna", "08:14", "00:00", 1),
    (23, "dyzur", "Chirurgia Ogólna", "00:00", "08:18", 0),
    (24, "work", "Urologia", "07:14", "15:00", 0),
    (25, "work", "Urologia", "07:06", "13:35", 0),
    (26, "work", "Chirurgia Onk.", "07:13", "16:16", 0),
    (30, "work", "Urologia", "07:29", "15:00", 0),
    (30, "dyzur", "Urologia", "15:00", "07:00", 1),
    (31, "work", "Urologia", "07:00", "09:50", 0),
]

# Expected (Godziny, Dyżury godz.) per day, straight off the photo.
EXPECTED = {d: (0.0, 0.0) for d in range(1, 32)}
EXPECTED.update({
    2: (8.23, 0.0), 3: (9.27, 0.0), 4: (8.18, 0.0), 5: (8.12, 0.0),
    6: (7.87, 17.18), 8: (0.0, 14.00), 9: (9.28, 0.0), 10: (8.00, 16.00),
    11: (3.85, 0.0), 12: (7.77, 16.00), 13: (10.18, 0.0), 14: (0.0, 15.53),
    15: (0.0, 8.77), 16: (7.90, 16.00), 17: (2.77, 0.0), 18: (8.07, 0.0),
    19: (7.72, 16.00), 20: (3.12, 0.0), 22: (0.0, 15.77), 23: (0.0, 8.30),
    24: (7.77, 0.0), 25: (6.48, 0.0), 26: (9.05, 0.0), 30: (7.52, 16.00),
    31: (2.83, 0.0),
})
EXPECTED_TOTAL = 303.52


def _dt(day, hhmm, offset=0):
    hh, mm = map(int, hhmm.split(":"))
    return datetime(YEAR, MONTH, day + offset, hh, mm, tzinfo=TZ)


USER = 111


@pytest.fixture
async def march_2026_workbook(tmp_path):
    for day, kind, oddzial, start, end, offset in RAW_ENTRIES:
        await db.add_manual_entry(USER, kind, oddzial, _dt(day, start), _dt(day, end, offset))
    entries = await db.get_entries_for_month(USER, YEAR, MONTH)
    out_path = tmp_path / "Grafik_2026_03.xlsx"
    generate_month_excel(entries, YEAR, MONTH, out_path)
    return openpyxl.load_workbook(out_path)


async def test_every_day_matches_photo(march_2026_workbook):
    ws = march_2026_workbook.active
    for day in range(1, 32):
        row = HEADER_ROW + day
        exp_godziny, exp_dyzur = EXPECTED[day]
        assert ws.cell(row=row, column=1).value == day, f"day label mismatch on row {row}"
        assert ws.cell(row=row, column=5).value == pytest.approx(exp_godziny, abs=0.005), f"Godziny day {day}"
        assert ws.cell(row=row, column=8).value == pytest.approx(exp_dyzur, abs=0.005), f"Dyżury godz. day {day}"
        assert ws.cell(row=row, column=9).value == f"=E{row}+H{row}"


async def test_department_labels(march_2026_workbook):
    ws = march_2026_workbook.active
    assert ws.cell(row=HEADER_ROW + 6, column=2).value == "Chirurgia Onk."
    assert ws.cell(row=HEADER_ROW + 14, column=2).value == "Chirurgia Ogólna"
    assert ws.cell(row=HEADER_ROW + 2, column=2).value == "Urologia"
    assert ws.cell(row=HEADER_ROW + 1, column=2).value in ("", None)


async def test_overnight_dyzur_recorded_on_start_day(march_2026_workbook):
    ws = march_2026_workbook.active
    row6 = HEADER_ROW + 6
    assert ws.cell(row=row6, column=6).value == "15:00"
    assert ws.cell(row=row6, column=7).value == "08:11"


async def test_monthly_total_matches_photo(march_2026_workbook):
    ws = march_2026_workbook.active
    sum_row = HEADER_ROW + 31 + 2
    assert ws.cell(row=sum_row, column=1).value == "SUMA GODZIN (miesiąc):"
    assert ws.cell(row=sum_row, column=9).value == pytest.approx(EXPECTED_TOTAL, abs=0.001)
